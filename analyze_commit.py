#!/usr/bin/env python3
"""
Script para analizar cambios en el commit 0fc6d57d9e52944629c05e2febb1aa0e6a870fba
y generar un archivo Excel con el análisis de problemas potenciales.
"""

import pandas as pd
import re
from typing import List, Dict, Tuple
from datetime import datetime

# Datos del commit obtenidos de GitHub API
COMMIT_DATA = {
    "sha": "0fc6d57d9e52944629c05e2febb1aa0e6a870fba",
    "message": "Simplify web demo with Vite",
    "files": [
        {
            "filename": "index.html",
            "additions": 29,
            "deletions": 39,
            "patch": """@@ -3,46 +3,36 @@
 <head>
   <meta charset="UTF-8" />
   <meta name="viewport" content="width=device-width, initial-scale=1.0" />
-  <meta name="theme-color" content="#111111" />
-  <title>TryOnMe – Discover Your Fit</title>
-  <link rel="stylesheet" href="css/style.css" />
-  <link rel="manifest" href="manifest.json" />
+  <title>TryOnMe</title>
+  <style>
+    body {
+      margin: 0;
+      padding: 0;
+      background: #000;
+      color: white;
+      font-family: 'Arial', sans-serif;
+      display: flex;
+      flex-direction: column;
+      min-height: 100vh;
+    }
+    #intro-video {
+      width: 100%;
+      height: auto;
+      max-height: 100vh;
+    }
+    footer {
+      margin-top: auto;
+      text-align: center;
+      padding: 1rem;
+      background-color: #000;
+      font-size: 0.9rem;
+      color: #ccc;
+    }
+  </style>
 </head>
 <body>
-  <a href="#main-content" class="skip-link">Saltar al contenido principal</a>
-  <header>
-    <img src="logo.webp" alt="TryOnMe Logo" class="logo" />
-    <nav aria-label="Navegación principal">
-      <ul>
-        <li><a href="#talla">Descubre tu talla</a></li>
-        <li><a href="#productos">Productos</a></li>
-        <li><a href="#contacto">Contacto</a></li>
-      </ul>
-    </nav>
-  </header>
-
-  <main id="main-content">
-    <section class="hero">
-      <video id="videoIntro" autoplay muted loop playsinline class="video-fondo">
-        <track kind="captions" src="captions_en.vtt" srclang="en" label="English captions">
-        Your browser does not support the video tag.
-      </video>
-      <div class="overlay">
-        <h1>Measure. Match. Live it.</h1>
-        <a href="https://tu-servidor.onrender.com/upload" class="cta-button">Descubre tu talla</a>
-      </div>
-    </section>
-
-    <section id="productos" class="productos">
-      <h2>Productos destacados</h2>
-      <div class="product-grid" id="product-grid"></div>
-    </section>
-
-  </main>
-
-  <footer>
-    <p>&copy; <span id="year">2025</span> TryOnMe. All rights reserved.</p>
-  </footer>
-  <script src="script.js"></script>
+  <video id="intro-video" autoplay muted loop playsinline></video>
+  <footer>© <span id="copyright-year"></span> TryOnMe. All rights reserved.</footer>
+  <script type="module" src="./script.js"></script>
 </body>
 </html>"""
        },
        {
            "filename": "package.json",
            "additions": 6,
            "deletions": 12,
            "patch": """@@ -1,19 +1,13 @@
 {
-  "name": "tryon-app",
+  "name": "tryonme",
   "version": "1.0.0",
   "scripts": {
-    "pretest": "npm install && pip install -r requirements.txt",
-    "test": "jest",
-    "deploy": "bash 04_Scripts_Produccion/ejecutar_agentes.sh",
-    "start": "node index.js",
-    "build": "node scripts/build.js"
+    "dev": "vite",
+    "build": "vite build",
+    "preview": "vite preview"
   },
   "dependencies": {},
   "devDependencies": {
-    "jest": "^30.0.4"
-  },
-  "engines": {
-    "node": ">=18"
-  },
-  "type": "module"
+    "vite": "^5.0.0"
+  }
 }"""
        },
        {
            "filename": "script.js",
            "additions": 46,
            "deletions": 45,
            "patch": """@@ -1,55 +1,56 @@
-const products = [
-  { name: "Classic Jacket", img: "jacket.webp" },
-  { name: "Urban Sneakers", img: "sneakers.webp" }
-];
+function safeDomExecution(callback) {
+  if (typeof window !== 'undefined' && typeof document !== 'undefined') {
+    document.addEventListener('DOMContentLoaded', callback);
+  }
+}
 
-async function loadCasaPavoReal() {
-  try {
-    const res = await fetch('products/casa_pavo_real.json');
-    if (res.ok) {
-      const collection = await res.json();
-      products.push(...collection);
-    }
-  } catch (err) {
-    console.error('Failed to load Casa Pavo Real collection', err);
+function getLocalizedVideo(lang = 'en') {
+  const videoMap = {
+    es: 'videos/intro_espanol.mp4',
+    fr: 'videos/intro_frances.mp4',
+    en: 'videos/intro_ingles.mp4'
+  };
+  const shortLang = lang.slice(0, 2).toLowerCase();
+  return videoMap[shortLang] || videoMap['en'];
+}
+
+function updateFooterYear() {
+  const yearEl = document.getElementById('copyright-year');
+  if (yearEl) yearEl.textContent = new Date().getFullYear();
+}
+
+function loadIntroVideo() {
+  const videoEl = document.getElementById('intro-video');
+  if (videoEl) {
+    videoEl.src = getLocalizedVideo(navigator.language);
+    videoEl.load();
   }
 }
 
+async function loadCasaPavoReal() {
+  console.log('🦚 Cargando datos de Casa Pavo Real...');
+}
+
 function displayProducts() {
-  const grid = document.getElementById('product-grid');
-  if (!grid) return;
-  products.forEach(p => {
-    const item = document.createElement('div');
-    item.className = 'product-item';
-    item.innerHTML = `<img src="${p.img}" alt="${p.name}"><p>${p.name}</p>`;
-    grid.appendChild(item);
-  });
+  console.log('🧥 Mostrando productos renderizados...');
 }
 
-if (typeof window !== 'undefined' && typeof document !== 'undefined') {
-  document.addEventListener('DOMContentLoaded', async () => {
+safeDomExecution(async () => {
+  try {
     await loadCasaPavoReal();
     displayProducts();
+    loadIntroVideo();
+    updateFooterYear();
+  } catch (error) {
+    console.error('🛑 Error en la ejecución de los superagentes:', error);
+  }
+});
 
-    const getLocalizedVideo = (lang = 'en') => {
-      const videoMap = {
-        es: 'videos/intro_espanol.mp4',
-        fr: 'videos/intro_frances.mp4',
-        en: 'videos/intro_ingles.mp4'
-      };
-      const matched = Object.keys(videoMap).find(key => lang.startsWith(key));
-      return videoMap[matched] || videoMap.en;
-    };
-
-    const lang = navigator.language || navigator.userLanguage || 'en';
-    const videoEl = document.getElementById('videoIntro');
-    if (videoEl) {
-      videoEl.src = getLocalizedVideo(lang);
-    }
-
-    const yearEl = document.getElementById('year');
-    if (yearEl) {
-      yearEl.textContent = new Date().getFullYear();
-    }
-  });
-}
+export {
+  safeDomExecution,
+  getLocalizedVideo,
+  updateFooterYear,
+  loadIntroVideo,
+  loadCasaPavoReal,
+  displayProducts
+};"""
        },
        {
            "filename": "vite.config.js",
            "additions": 7,
            "deletions": 0,
            "patch": """@@ -0,0 +1,7 @@
+export default {
+  root: '.',
+  server: {
+    open: true,
+    port: 3000
+  }
+};"""
        }
    ]
}


def parse_diff_lines(patch: str) -> List[Tuple[str, str, int]]:
    """
    Parsea las líneas del diff y extrae los cambios específicos.
    Retorna lista de tuplas: (tipo_cambio, contenido, numero_linea)
    """
    lines = patch.split('\n')
    changes = []
    current_line = 0
    
    for line in lines:
        if line.startswith('@@'):
            # Extraer número de línea inicial del diff
            match = re.search(r'\+(\d+)', line)
            if match:
                current_line = int(match.group(1))
        elif line.startswith('+') and not line.startswith('+++'):
            changes.append(('added', line[1:], current_line))
            current_line += 1
        elif line.startswith('-') and not line.startswith('---'):
            changes.append(('removed', line[1:], current_line))
        elif not line.startswith('\\'):
            current_line += 1
    
    return changes


def analyze_html_issues(changes: List[Tuple[str, str, int]]) -> List[Dict]:
    """Analiza problemas específicos en cambios de HTML."""
    issues = []
    
    for change_type, content, line_num in changes:
        content = content.strip()
        
        if change_type == 'removed':
            if 'meta name="theme-color"' in content:
                issues.append({
                    'file': 'index.html',
                    'lines': f'Línea {line_num}',
                    'problem': 'Se eliminó la meta etiqueta theme-color que define el color de la barra de estado en móviles',
                    'solution': 'Agregar <meta name="theme-color" content="#000000"> para mantener consistencia visual en dispositivos móviles'
                })
            
            elif 'link rel="stylesheet"' in content:
                issues.append({
                    'file': 'index.html',
                    'lines': f'Línea {line_num}',
                    'problem': 'Se eliminó la referencia al archivo CSS externo, perdiendo estilos organizados',
                    'solution': 'Mantener el archivo CSS externo para mejor organización o asegurar que todos los estilos críticos estén inline'
                })
            
            elif 'link rel="manifest"' in content:
                issues.append({
                    'file': 'index.html',
                    'lines': f'Línea {line_num}',
                    'problem': 'Se eliminó el manifest.json, perdiendo capacidades PWA (Progressive Web App)',
                    'solution': 'Mantener <link rel="manifest" href="manifest.json"> para conservar funcionalidades PWA como instalación y offline'
                })
            
            elif 'skip-link' in content:
                issues.append({
                    'file': 'index.html',
                    'lines': f'Línea {line_num}',
                    'problem': 'Se eliminó el enlace de salto de accesibilidad, reduciendo usabilidad para usuarios con discapacidades',
                    'solution': 'Mantener <a href="#main-content" class="skip-link">Saltar al contenido principal</a> para cumplir estándares WCAG'
                })
            
            elif '<header>' in content or '<nav>' in content:
                issues.append({
                    'file': 'index.html',
                    'lines': f'Línea {line_num}',
                    'problem': 'Se eliminó la navegación principal, perdiendo funcionalidad de navegación del sitio',
                    'solution': 'Mantener elementos de navegación o implementar una navegación alternativa para acceso a secciones'
                })
        
        elif change_type == 'added':
            if 'style>' in content:
                issues.append({
                    'file': 'index.html',
                    'lines': f'Línea {line_num}',
                    'problem': 'Se agregaron estilos inline que pueden ser difíciles de mantener y reutilizar',
                    'solution': 'Considerar mover estilos a un archivo CSS externo para mejor organización y reutilización'
                })
            
            if 'script type="module"' in content:
                issues.append({
                    'file': 'index.html',
                    'lines': f'Línea {line_num}',
                    'problem': 'El uso de ES modules puede no ser compatible con navegadores antiguos',
                    'solution': 'Agregar polyfills o bundling para garantizar compatibilidad con navegadores legacy'
                })
    
    return issues


def analyze_package_json_issues(changes: List[Tuple[str, str, int]]) -> List[Dict]:
    """Analiza problemas específicos en cambios de package.json."""
    issues = []
    
    for change_type, content, line_num in changes:
        content = content.strip()
        
        if change_type == 'removed':
            if '"test":' in content:
                issues.append({
                    'file': 'package.json',
                    'lines': f'Línea {line_num}',
                    'problem': 'Se eliminó el script de testing, perdiendo capacidad de ejecutar pruebas automatizadas',
                    'solution': 'Agregar script de testing compatible con Vite: "test": "vitest" y instalar vitest como dependencia'
                })
            
            elif '"deploy":' in content:
                issues.append({
                    'file': 'package.json',
                    'lines': f'Línea {line_num}',
                    'problem': 'Se eliminó el script de deploy automatizado, perdiendo proceso de despliegue',
                    'solution': 'Adaptar el script de deploy para trabajar con Vite: "deploy": "vite build && bash 04_Scripts_Produccion/ejecutar_agentes.sh"'
                })
            
            elif '"engines":' in content:
                issues.append({
                    'file': 'package.json',
                    'lines': f'Línea {line_num}',
                    'problem': 'Se eliminó la especificación de versión de Node.js, perdiendo control de compatibilidad',
                    'solution': 'Mantener "engines": {"node": ">=18"} para garantizar compatibilidad y evitar errores en producción'
                })
            
            elif '"type": "module"' in content:
                issues.append({
                    'file': 'package.json',
                    'lines': f'Línea {line_num}',
                    'problem': 'Se eliminó "type": "module", lo que puede causar problemas con imports ES6',
                    'solution': 'Mantener "type": "module" o configurar Vite correctamente para manejar módulos ES6'
                })
        
        elif change_type == 'added':
            if '"vite":' in content:
                issues.append({
                    'file': 'package.json',
                    'lines': f'Línea {line_num}',
                    'problem': 'Dependencia de Vite sin versión específica puede causar problemas de compatibilidad',
                    'solution': 'Especificar versión exacta de Vite y agregar dependencias necesarias como @vitejs/plugin-* según necesidades'
                })
    
    return issues


def analyze_javascript_issues(changes: List[Tuple[str, str, int]]) -> List[Dict]:
    """Analiza problemas específicos en cambios de JavaScript."""
    issues = []
    
    for change_type, content, line_num in changes:
        content = content.strip()
        
        if change_type == 'removed':
            if 'const products = [' in content:
                issues.append({
                    'file': 'script.js',
                    'lines': f'Línea {line_num}',
                    'problem': 'Se eliminó la definición inicial de productos, perdiendo datos de fallback',
                    'solution': 'Mantener productos por defecto o implementar un mecanismo de fallback cuando falle la carga de datos externos'
                })
            
            elif 'fetch(' in content:
                issues.append({
                    'file': 'script.js',
                    'lines': f'Línea {line_num}',
                    'problem': 'Se eliminó la funcionalidad de carga dinámica de productos, perdiendo contenido dinámico',
                    'solution': 'Reimplementar la carga de productos desde casa_pavo_real.json o proporcionar datos alternativos'
                })
            
            elif 'getElementById(' in content and 'appendChild(' in content:
                issues.append({
                    'file': 'script.js',
                    'lines': f'Línea {line_num}',
                    'problem': 'Se eliminó la lógica de renderizado de productos en el DOM',
                    'solution': 'Reimplementar la función de renderizado de productos para mostrar contenido dinámico en la página'
                })
        
        elif change_type == 'added':
            if 'console.log(' in content and ('🦚' in content or '🧥' in content):
                issues.append({
                    'file': 'script.js',
                    'lines': f'Línea {line_num}',
                    'problem': 'Se agregaron console.log con emojis que no implementan funcionalidad real',
                    'solution': 'Reemplazar console.log con implementación real de las funciones loadCasaPavoReal y displayProducts'
                })
            
            if 'export {' in content:
                issues.append({
                    'file': 'script.js',
                    'lines': f'Línea {line_num}',
                    'problem': 'Se agregaron exports sin un sistema de módulos configurado correctamente',
                    'solution': 'Asegurar que el bundler (Vite) esté configurado para manejar ES modules o cambiar a CommonJS'
                })
            
            if 'navigator.language' in content:
                issues.append({
                    'file': 'script.js',
                    'lines': f'Línea {line_num}',
                    'problem': 'Uso de navigator.language sin validación puede fallar en algunos navegadores',
                    'solution': 'Agregar fallback: navigator.language || navigator.userLanguage || "en" para mayor compatibilidad'
                })
    
    return issues


def analyze_vite_config_issues(changes: List[Tuple[str, str, int]]) -> List[Dict]:
    """Analiza problemas específicos en la configuración de Vite."""
    issues = []
    
    for change_type, content, line_num in changes:
        content = content.strip()
        
        if change_type == 'added':
            if 'export default {' in content:
                issues.append({
                    'file': 'vite.config.js',
                    'lines': f'Línea {line_num}',
                    'problem': 'Configuración básica de Vite sin optimizaciones ni plugins necesarios',
                    'solution': 'Agregar plugins para HTML, CSS processing y optimizaciones: import { defineConfig } from "vite" y configurar build options'
                })
            
            if 'port: 3000' in content:
                issues.append({
                    'file': 'vite.config.js',
                    'lines': f'Línea {line_num}',
                    'problem': 'Puerto fijo puede causar conflictos si está ocupado',
                    'solution': 'Configurar puerto dinámico: port: process.env.PORT || 3000, o usar strictPort: false'
                })
            
            if 'open: true' in content:
                issues.append({
                    'file': 'vite.config.js',
                    'lines': f'Línea {line_num}',
                    'problem': 'Auto-apertura del navegador puede ser molesta en entornos de desarrollo',
                    'solution': 'Hacer configurable: open: process.env.NODE_ENV === "development" && !process.env.CI'
                })
    
    return issues


def generate_excel_report():
    """Genera el reporte Excel con el análisis de cambios."""
    all_issues = []
    
    # Analizar cada archivo
    for file_data in COMMIT_DATA['files']:
        filename = file_data['filename']
        patch = file_data['patch']
        
        # Saltear package-lock.json como se especifica
        if filename == 'package-lock.json':
            continue
        
        changes = parse_diff_lines(patch)
        
        if filename == 'index.html':
            issues = analyze_html_issues(changes)
        elif filename == 'package.json':
            issues = analyze_package_json_issues(changes)
        elif filename == 'script.js':
            issues = analyze_javascript_issues(changes)
        elif filename == 'vite.config.js':
            issues = analyze_vite_config_issues(changes)
        else:
            issues = []
        
        all_issues.extend(issues)
    
    # Crear DataFrame
    df = pd.DataFrame(all_issues, columns=[
        'file', 'lines', 'problem', 'solution'
    ])
    
    # Renombrar columnas al español
    df.columns = [
        'Archivo afectado',
        'Línea(s) modificadas', 
        'Descripción del problema',
        'Propuesta de solución'
    ]
    
    # Generar nombre de archivo con timestamp
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'analisis_commit_{COMMIT_DATA["sha"][:8]}_{timestamp}.xlsx'
    
    # Crear archivo Excel con formato
    with pd.ExcelWriter(filename, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Análisis de Cambios', index=False)
        
        # Obtener el workbook y worksheet para formatear
        workbook = writer.book
        worksheet = writer.sheets['Análisis de Cambios']
        
        # Ajustar ancho de columnas
        worksheet.column_dimensions['A'].width = 20  # Archivo afectado
        worksheet.column_dimensions['B'].width = 25  # Línea(s) modificadas
        worksheet.column_dimensions['C'].width = 50  # Descripción del problema
        worksheet.column_dimensions['D'].width = 60  # Propuesta de solución
        
        # Aplicar formato a los headers
        from openpyxl.styles import Font, PatternFill, Alignment
        
        header_font = Font(bold=True, color='FFFFFF')
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Aplicar word wrap a todas las celdas
        for row in worksheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    print(f"✅ Archivo Excel generado: {filename}")
    print(f"📊 Total de problemas identificados: {len(all_issues)}")
    print(f"📋 Commit analizado: {COMMIT_DATA['sha']}")
    print(f"💬 Mensaje del commit: {COMMIT_DATA['message']}")
    
    return filename


if __name__ == '__main__':
    generate_excel_report()